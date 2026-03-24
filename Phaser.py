import openpyxl as xl
from constants import LOS_C_THRESHOLD, LOS_D_THRESHOLD, MAX_LANES_PER_DIRECTION
from b_optimization import b_optimization
from c_optimization import c_optimization
from queue_length import queue_length

from personal_filter import personal_filter
from rakal_capacity import rakal_capacity
import timeit
import platform
import shutil
# import pulp as pl
# import os
# import pprint
# import math


# לשאלות ותמיכה טכנית ניתן לפנות לעמרי מטר omrimatar@gmail.com

def main(queue_params=None):
    cwd = 'os.getcwd()'  # solverdir = r'cbc-windeps-win64-msvc16-mtd\bin\cbc.exe'  # extracted and renamed CBC solver binary
    # solverdir = os.path.join(cwd, solverdir)
    # solver = pl.COIN_CMD(solverdir)

    if queue_params is None:
        queue_params = {}
    solver = 1
    start = timeit.default_timer()
    run = 0
    junc_list = []
    _car_sum_am = _car_sum_pm = _pulp_vars_am = _pulp_vars_pm = None
    excel_properties_list= []
    queue_max_list = [0] * 28
    # poisson_cars = [0] * 12
    runlist = ["Morning","Evening"]
    while run < 2:
        junc_list.append(runlist[run])
        if run == 0:
            print("             morning:")
        else:
            print("             night:")

        def write_to_excel(v_over_c, sum_of_images, pulp_vars, run, real_capacity):

            wb2 = xl.load_workbook('OUTPUT.xlsx')
            ws2 = wb2.active
            write_names = [''] * 150
            write_values = [''] * 150
            if run == 0:
                namesA = 'A'
                valuesB = 'B'
            else:
                namesA = 'E'
                valuesB = 'F'

            for m in range(len(pulp_vars)):
                write_names[m] = list(pulp_vars.keys())[len(pulp_vars) - m - 1]
                write_values[m] = list(pulp_vars.values())[len(pulp_vars) - m - 1]

            ws2['A1'] = 'morning'
            ws2['E1'] = 'night'
            ws2[namesA + '2'] = 'capacity'
            ws2[valuesB + '2'] = real_capacity
            ws2[namesA + '3'] = 'v/C'
            ws2[valuesB + '3'] = v_over_c
            ws2[namesA + '4'] = 'sum of images'
            ws2[valuesB + '4'] = sum_of_images
            ws2[namesA + '5'] = 'names'
            ws2[valuesB + '5'] = 'values'

            for i in range(150):
                row = str(i + 6)
                ws2[namesA + row] = write_names[i]
                ws2[valuesB + row] = write_values[i]
            if platform.system() == 'Windows':
                import ctypes
                MessageBox = ctypes.windll.user32.MessageBoxW
                MB_RETRYCANCEL = 5
                IDRETRY = 4
                while True:
                    try:
                        wb2.save('OUTPUT.xlsx')
                        break
                    except PermissionError:
                        result = MessageBox(None,
                                            'OUTPUT.xlsx is open in Excel.\nPlease close it and press Retry.',
                                            'Phaser error', MB_RETRYCANCEL)
                        if result != IDRETRY:
                            raise PermissionError('OUTPUT.xlsx is locked — please close it.')
            else:
                wb2.save('OUTPUT.xlsx')

            wb2.close()

        def print_dict(d):
            for item in d.items():
                print(item[0], '=', item[1])

        def suppress_null(val):
            if not val:
                return 0
            else:
                return val

        def read_from_excel(run):
            wb = xl.load_workbook('volume_calculator.xlsx', data_only=True)
            ws = wb.active
            volume = []
            lanes = []
            nataz = []
            junc_nataz=[]
            streets = []


            junc_instructions = [suppress_null(ws.cell(row=36 + i, column=19).value) for i in range(5)]

            for i in range(4):
                l = [suppress_null(ws.cell(row=4 + run, column=4 + 4 * i + j).value) for j in range(3)]
                volume += l
                l = [suppress_null(ws.cell(row=8, column=3 + 8 * i + j).value) for j in range(7)]
                lanes += l
                l = [suppress_null(ws.cell(row=9, column=3 + 8 * i + j).value) for j in range(7)]
                nataz += l
                l = [suppress_null(ws.cell(row=9, column=3 + 8 * i + j).value) for j in range(7)]
                junc_nataz += l

            instructions = [suppress_null(ws.cell(row=36 + i, column=22).value) for i in range(11)]

            # ---- Inflation factor check (cell V46) ----
            _raw_inflation = ws.cell(row=46, column=22).value  # read before suppress_null
            if _raw_inflation is None:
                # blank cell: default to 1.0 (no adjustment)
                instructions[10] = 1.0
            else:
                _inflation = _raw_inflation
                if _inflation <= 0:
                    raise ValueError(
                        f"Invalid inflation factor in cell V46: value = {_inflation}. "
                        f"The inflation factor must be a positive number greater than 0 "
                        f"(use 1.0 for no adjustment, e.g. 1.1 for +10%). "
                        f"A value of 0 or negative would zero out or invert all traffic volumes."
                    )
                instructions[10] = _inflation

            # ---- Apply inflation and validate volume cells ----
            _vol_cells = ['D4','E4','F4','H4','I4','J4','L4','M4','N4','P4','Q4','R4']
            _vol_names = ['North right','North through','North left',
                          'South right','South through','South left',
                          'East right', 'East through', 'East left',
                          'West right', 'West through', 'West left']
            for i in range(12):
                try:
                    volume[i] = round(volume[i] * instructions[10], 0)
                except Exception:
                    _row = 4 + run
                    raise ValueError(
                        f"Non-numeric volume value in {_vol_names[i]} "
                        f"(cell {_vol_cells[i][0]}{_row}). "
                        f"All traffic volume cells in rows 4-5 must contain numbers. "
                        f"Check the range D4:R5."
                    )
                if volume[i] < 0:
                    _row = 4 + run
                    raise ValueError(
                        f"Negative volume after inflation in {_vol_names[i]} "
                        f"(cell {_vol_cells[i][0]}{_row}): value = {int(volume[i])}. "
                        f"Check the raw volume (rows 4-5) and the inflation factor (cell V46). "
                        f"Both must be non-negative."
                    )

            rakal_instructions = [suppress_null(ws.cell(row=36 + i, column=26).value) for i in range(6)]
            _rakal_keys = ['lrt_enabled (Z36)', 'cycle_time (Z37)', 'lost_time (Z38)',
                           'headway (Z39)', 'mcu (Z40)', 'gen_lost_time (Z41)']
            for i in range(6):
                if i == 4 and rakal_instructions[i] == 1.125: i = i + 1
                if isinstance(rakal_instructions[i], int) == False:
                    raise ValueError(
                        f"Non-integer value in LRT parameters table, field {_rakal_keys[i]}: "
                        f"value = {rakal_instructions[i]!r}. "
                        f"All values in column Z rows 36-41 must be integers "
                        f"(exception: mcu in Z40 may be 1.125)."
                    )
            _instr_keys = ['capacity (V36)', 'nlsl (V37)', 'elwl (V38)', 'img5 (V39)',
                           'img6 (V40)', 'geo_ns (V41)', 'geo_ew (V42)', 'optimize (V43)',
                           'lrt_orig_ns (V44)', 'lrt_orig_ew (V45)', 'inflation (V46)']
            for i in range(11):
                # אם בעתיד מגדילים את הריינג' לשנות את פקדות הברייק שמתחת
                if i == 10 and isinstance(instructions[i], float) == True: break
                if isinstance(instructions[i], int) == False:
                    raise ValueError(
                        f"Non-integer value in parameters table, field {_instr_keys[i]}: "
                        f"value = {instructions[i]!r}. "
                        f"All values in column V rows 36-45 must be integers "
                        f"(V46 inflation may be a decimal)."
                    )

            # instructionscheck = [s for s in rakal_instructions if s.isdigit()]

            #        x=instructions[i]+1
            #        y=rakal_instructionsinstructions[i]+1
            # sum [instructions,0]
            # sum [rakal_instructions,0]
            # except:

            for r in range(12):
                if (r + 2) % 3 != 0:
                    volume[r] = round(volume[r] * rakal_instructions[4], 0)
            if instructions[7] == 1:
                for m in range(28):
                    nataz[m] = 0
            for s in range(4):
                l = [suppress_null(ws.cell(row=4, column=22 + s).value)]
                if l == 0: l = ""
                streets += l
            # print("streets=", streets)

            wb.properties
            return volume, lanes, instructions, rakal_instructions, nataz, streets, junc_instructions,wb.properties,junc_nataz


        def update_excel(optimal_lanes, min_v_c):
            shutil.copy2('volume_calculator.xlsx', 'volume_calculator_backup.xlsx')
            for i in range(4):
                for j in range(7):
                    ws.cell(row=8, column=3 + 8 * i + j).value = optimal_lanes[i * 7 + j]
            ws['M31'] = min_v_c

            wb.save('volume_calculator.xlsx')
            wb.close()

        wb = xl.load_workbook('volume_calculator.xlsx')
        ws = wb.active
        print("after morning")
        volume, lanes, instructions, rakal_instructions, nataz, streets, junc_instructions, excel_properties,junc_nataz = read_from_excel(run)

        # שליחת נתונים לגאנק
        junc_list.append(volume)
        junc_list.append(lanes)
        junc_list.append(instructions)
        junc_list.append(rakal_instructions)
        junc_list.append(junc_nataz)

        NcountR = round(volume[0], 0)
        NcountT = round(volume[1], 0)
        NcountL = round(volume[2], 0)
        ScountR = round(volume[3], 0)
        ScountT = round(volume[4], 0)
        ScountL = round(volume[5], 0)
        EcountR = round(volume[6], 0)
        EcountT = round(volume[7], 0)
        EcountL = round(volume[8], 0)
        WcountR = round(volume[9], 0)
        WcountT = round(volume[10], 0)
        WcountL = round(volume[11], 0)

        if instructions[7] == 1:
            run = 2

            # רקורסיה ליצירת מערכים אפשריים של צפון דרום
            def zeros(num):
                return [0 for i in range(num)]

            def generate(n, size):
                if (size == 1):
                    return [[n]]
                elif (n == 0):
                    return [zeros(size)]
                else:
                    elements1 = [list(map(lambda lst: [i] + lst, generate(n - i, size - 1))) for i in range(n + 1)]
                    return [j for i in elements1 for j in i]

            def filterPredicate(lst):

                Nr = lst[0]
                Nrt = lst[1]
                Nt = lst[2]
                Ntl = lst[3]
                Nl = lst[4]
                Nrtl = lst[5]
                Nrl = lst[6]

                Sr = lst[7]
                Srt = lst[8]
                St = lst[9]
                Stl = lst[10]
                Sl = lst[11]
                Srtl = lst[12]
                Srl = lst[13]
                # סינון המערך לאפשרויות מתאימות בלבד. יש להחליף בעתיד את הערכים 50 בערכים שאובים מאקסל.

                if Nr + Nrt + Nt + Ntl + Nl + Nrtl + Nrl <= MAX_LANES_PER_DIRECTION \
                        and Sr + Srt + St + Stl + Sl + Srtl + Srl <= MAX_LANES_PER_DIRECTION \
                        and Nrt + Nrl + Nrtl <= 1 \
                        and Srt + Srl + Srtl <= 1 \
                        and Ntl + Nrl + Nrtl <= 1 \
                        and Stl + Srl + Srtl <= 1 \
                        and Nt + 100 * Nrl <= 100 \
                        and St + 100 * Srl <= 100 \
                        and Nrt + Ntl + Nrl + Nt + 100 * Nrtl <= 100 \
                        and Srt + Stl + Srl + St + 100 * Srtl <= 100 \
                        and NcountR >= Nr + Nrt + Nrtl + Nrl \
                        and NcountT >= Nt + Nrt + Nrtl + Ntl \
                        and NcountL >= Nl + Nrl + Nrtl + Ntl \
                        and ScountR >= Sr + Srt + Srtl + Srl \
                        and ScountT >= St + Srt + Srtl + Stl \
                        and ScountL >= Sl + Srl + Srtl + Stl \
                        and NcountR <= (Nr + Nrt + Nrtl + Nrl) * 10000 \
                        and NcountT <= (Nt + Nrt + Nrtl + Ntl) * 10000 \
                        and NcountL <= (Nl + Ntl + Nrtl + Nrl) * 10000 \
                        and ScountR <= (Sr + Srt + Srtl + Srl) * 10000 \
                        and ScountT <= (St + Srt + Srtl + Stl) * 10000 \
                        and ScountL <= (Sl + Stl + Srtl + Srl) * 10000 \
                        and 1:
                    return True

            # רקורסיה ליצירת מערכים אפשריים של מזרח מערב
            def zeros(num):
                return [0 for i in range(num)]

            def generate(n, size):
                if (size == 1):
                    return [[n]]
                elif (n == 0):
                    return [zeros(size)]
                else:
                    elements1 = [list(map(lambda lst: [i] + lst, generate(n - i, size - 1))) for i in range(n + 1)]
                    return [j for i in elements1 for j in i]

            def filterPredicate2(lst):

                Er = lst[0]
                Ert = lst[1]
                Et = lst[2]
                Etl = lst[3]
                El = lst[4]
                Ertl = lst[5]
                Erl = lst[6]

                Wr = lst[7]
                Wrt = lst[8]
                Wt = lst[9]
                Wtl = lst[10]
                Wl = lst[11]
                Wrtl = lst[12]
                Wrl = lst[13]
                # סינון המערך לאפשרויות מתאימות בלבד. יש להחליף בעתיד את הערכים 50 בערכים שאובים מאקסל.

                if Er + Ert + Et + Etl + El + Ertl + Erl <= MAX_LANES_PER_DIRECTION \
                        and Wr + Wrt + Wt + Wtl + Wl + Wrtl + Wrl <= MAX_LANES_PER_DIRECTION \
                        and Ert + Erl + Ertl <= 1 \
                        and Wrt + Wrl + Wrtl <= 1 \
                        and Etl + Erl + Ertl <= 1 \
                        and Wtl + Wrl + Wrtl <= 1 \
                        and Et + 100 * Erl <= 100 \
                        and Wt + 100 * Wrl <= 100 \
                        and Ert + Etl + Erl + Et + 100 * Ertl <= 100 \
                        and Wrt + Wtl + Wrl + Wt + 100 * Wrtl <= 100 \
                        and EcountR >= Er + Ert + Ertl + Erl \
                        and EcountT >= Et + Ert + Ertl + Etl \
                        and EcountL >= El + Erl + Ertl + Etl \
                        and WcountR >= Wr + Wrt + Wrtl + Wrl \
                        and WcountT >= Wt + Wrt + Wrtl + Wtl \
                        and WcountL >= Wl + Wrl + Wrtl + Wtl \
                        and EcountR <= (Er + Ert + Ertl + Erl) * 10000 \
                        and EcountT <= (Et + Ert + Ertl + Etl) * 10000 \
                        and EcountL <= (El + Etl + Ertl + Erl) * 10000 \
                        and WcountR <= (Wr + Wrt + Wrtl + Wrl) * 10000 \
                        and WcountT <= (Wt + Wrt + Wrtl + Wtl) * 10000 \
                        and WcountL <= (Wl + Wtl + Wrtl + Wrl) * 10000 \
                        and 1:
                    return True

            def combinList(h2, h3):
                combinations = []
                for x in h2:
                    for y in h3:
                        combinations.append(x + y)
                return combinations

            a = instructions[5]
            n = 14
            h = generate(a, n)
            h2 = list(filter(filterPredicate, h))
            print(h2)
            b = instructions[6]
            n = 14
            h = generate(b, n)
            h3 = list(filter(filterPredicate2, h))
            print(h3)

            combinations2 = combinList(h2, h3)
            filtered = list(filter(personal_filter, combinations2))
            min_v_c = 100.0
            for l in filtered:
                car_sum = b_optimization(volume, l, nataz, solver)
                v_c, sum_of_images, current_pulp_vars, images_values, lrt_compatible_phases = c_optimization(car_sum, instructions, nataz,
                                                                                      solver)
                if v_c < min_v_c:
                    min_v_c = v_c
                    optimal_lanes = l
                    pulp_vars = current_pulp_vars
                print(v_c, l)

            print('\n\n\noptimal lanes =', optimal_lanes)
            print('v/c =', min_v_c)
            print_dict(pulp_vars)
            update_excel(optimal_lanes, min_v_c)

        else:
            # בדיקת צומת יחיד על סמך הניתוב שהוזן
            real_capacity = instructions[0]
            car_sum = b_optimization(volume, lanes, nataz, solver)
            v_c, sum_of_images, current_pulp_vars, images_values, lrt_compatible_phases = c_optimization(car_sum, instructions, nataz, solver)
            queue_list = queue_length(car_sum, current_pulp_vars, **queue_params)
            if run == 0:
                _car_sum_am = list(car_sum)
                _pulp_vars_am = dict(current_pulp_vars)
            else:
                _car_sum_pm = list(car_sum)
                _pulp_vars_pm = dict(current_pulp_vars)
            if rakal_instructions[0] == 1:
                v_c, real_capacity = rakal_capacity(instructions, rakal_instructions, images_values, lrt_compatible_phases)
            write_to_excel(v_c, sum_of_images, current_pulp_vars, run, real_capacity)
            for keys, values in current_pulp_vars.items():
                if values > 0:
                    print(keys, '=', values)
            # print_dict(current_pulp_vars)

            if v_c < LOS_C_THRESHOLD:
                LOS = "C"
            elif v_c < LOS_D_THRESHOLD:
                LOS = "D"
            elif v_c < 1:
                LOS = "E"
            else:
                LOS = "F"

            if run == 0:
                morning_LOS = LOS
            else:
                night_LOS = LOS

            for i in range(28):
                if queue_list[i] > queue_max_list[i]:
                    queue_max_list[i] = queue_list[i]

            # counter=0
            # for i in range(12):
            # counter+=0.333
            # ilanes= math.floor(counter)
            # l = poisson_dist(volume[i],lanes[2*i+ilanes])
            # if l>poisson_cars[i]:
            # poisson_cars[i]= l
            # print (poisson_cars)

            # שליחת נתונים לגאנק
            junc_list.append(v_c)
            junc_list.append(sum_of_images)
            junc_list.append(current_pulp_vars)
            junc_list.append(real_capacity)
            if run==1:
                junc_list.append(streets)
                junc_list.append(junc_instructions)
                excel_properties_list.append(excel_properties)

        run = run + 1
        print("junc_list")
        print (excel_properties_list)
        for li in junc_list:
            print(li)
            # volume, lanes, instructions, rakal_instructions, nataz, v_c, sum_of_images, current_pulp_vars, run, real_capacity =

    # stop = timeit.default_timer()
    # print('Time:', stop - start)

    # name= ["NLlength","SLlength","ELlength","WLlength"]
    name = ["NR_length", "NRT_length", "NT_length", "NTL_length", "NL_length", "NRTL_length", "NRL_length", "SR_length",
            "SRT_length", "ST_length", "STL_length", "SL_length", "SRTL_length", "SRL_length",
            "ER_length", "ERT_length", "ET_length", "ETL_length", "EL_length", "ERTL_length", "ERL_length", "WR_length",
            "WRT_length", "WT_length", "WTL_length", "WL_length", "WRTL_length", "WRL_length"]
    car_length_dict = dict(zip(name, queue_max_list))
    for i in (name):
        if car_length_dict[i] > 0:
            print(i, "=", car_length_dict[i])
    # pprint.pprint (car_length_dict)

    #MessageBox = ctypes.windll.user32.MessageBoxW
    #MessageBox(None, 'calculation complete \n\n\nAM LOS=' + morning_LOS + '\nPM LOS=' + night_LOS, 'Phaser massage', 0)

    # for i in range(2):
    #    print ('i = ', i)

    extra_data = {
        "car_sum_am":    _car_sum_am,
        "car_sum_pm":    _car_sum_pm,
        "pulp_vars_am":  _pulp_vars_am,
        "pulp_vars_pm":  _pulp_vars_pm,
    }
    return (junc_list, excel_properties_list, car_length_dict, extra_data)


if __name__ == "__main__":
    main()
