"""
add_hcm_sheets.py
Run once to add three HCM 7th Edition sheets to the existing volume_calculator.xlsx.

New sheets added:
  • HCM Input      — new inputs only (green times, T/k/PF/I/l) + live references
                     to volumes, lanes, capacity, cycle time from the main sheet
  • HCM Delay      — d1·PF + d2, LOS per approach × period
  • HCM 95th Queue — Nq1+Nq2, Q95 in vehicles and metres
"""

import shutil
import os
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── Source file ───────────────────────────────────────────────────────────────
SRC = os.path.join(os.path.dirname(__file__), "volume_calculator.xlsx")
MAIN = "מחשבון נפח קובע"   # existing sheet name

def _ref(cell: str) -> str:
    """Build a cross-sheet reference string like ='שם גיליון'!A1"""
    return f"='{MAIN}'!{cell}"

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY    = "0D1B2A"
C_STEEL   = "1B3A5C"
C_DIRHEAD = "2C3E6B"
C_INPUT   = "FFFCE8"
C_REF     = "EBF5FB"   # blue-tint for auto-linked cells
C_CALC    = "F0F0F8"
C_GREY    = "F5F5F5"
C_DKGREY  = "666666"

LOS_STYLE = {
    "A": ("C6EFCE", "1E7A47"),
    "B": ("DDEEFF", "1A5276"),
    "C": ("FFF9C4", "7D6608"),
    "D": ("FFE0B2", "B15500"),
    "E": ("FFCDD2", "B71C1C"),
    "F": ("CC2200", "FFFFFF"),
}

DIR_BG   = {"N": "E8F4FD", "S": "E8F5E9", "E": "FFF8E1", "W": "FCE4EC"}
DIRS     = ["N", "S", "E", "W"]
DIR_EN   = {"N": "North", "S": "South", "E": "East", "W": "West"}

# Column index (1-based) per direction in the HCM Input / result sheets
DIR_COL_N = {"N": 2, "S": 3, "E": 4, "W": 5}

# ── Main-sheet cell references for volumes and lanes ─────────────────────────
# Morning volumes (row 4):  N=D4+E4+F4, S=H4+I4+J4, E=L4+M4+N4, W=P4+Q4+R4
# Evening volumes (row 5):  same cols, row 5
# Lanes (row 8):  N=C8:I8, S=K8:Q8, E=S8:Y8, W=AA8:AG8

_VOL_CELLS = {
    "N": ("D", "E", "F"),
    "S": ("H", "I", "J"),
    "E": ("L", "M", "N"),
    "W": ("P", "Q", "R"),
}
_LANE_RANGES = {
    "N": "C8:I8",
    "S": "K8:Q8",
    "E": "S8:Y8",
    "W": "AA8:AG8",
}

def _vol_ref(d: str, row: int) -> str:
    """Formula: sum of R+T+L volumes for direction d, row = 4 (AM) or 5 (PM)."""
    cols = _VOL_CELLS[d]
    return "=" + "+".join(f"'{MAIN}'!{c}{row}" for c in cols)

def _lane_ref(d: str) -> str:
    return f"=SUM('{MAIN}'!{_LANE_RANGES[d]})"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _border(color="CCCCCC"):
    s = _side(color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, italic=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAPC  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_THIN   = _border()

def _hdr_cell(ws, row, col, text, bg=C_NAVY, size=10, wrap=False):
    c = ws.cell(row, col, text)
    c.font      = _font(size=size, bold=True, color="FFFFFF")
    c.fill      = _fill(bg)
    c.alignment = _WRAPC if wrap else _CENTER
    c.border    = _border(color="33557A")
    return c

def _section_hdr(ws, row, text, last_col, bg=C_STEEL):
    ws.merge_cells(f"A{row}:{get_column_letter(last_col)}{row}")
    c = ws.cell(row, 1, f"  {text}")
    c.font      = _font(size=10, bold=True, color="FFFFFF")
    c.fill      = _fill(bg)
    c.alignment = _LEFT
    ws.row_dimensions[row].height = 17

def _input_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row, col, value)
    c.font      = _font(bold=True, color="1A3A6B")
    c.fill      = _fill(C_INPUT)
    c.alignment = _CENTER
    c.border    = _THIN
    if fmt:
        c.number_format = fmt
    return c

def _ref_cell(ws, row, col, formula, fmt="0", bold=False):
    """Auto-linked cell (blue tint, read-only visually)."""
    c = ws.cell(row, col, formula)
    c.font      = _font(bold=bold, color="1A5276")
    c.fill      = _fill(C_REF)
    c.alignment = _CENTER
    c.border    = _THIN
    c.number_format = fmt
    return c

def _calc_cell(ws, row, col, formula, fmt="0.0", bold=False):
    c = ws.cell(row, col, formula)
    c.font      = _font(bold=bold)
    c.fill      = _fill(C_CALC)
    c.alignment = _CENTER
    c.border    = _THIN
    c.number_format = fmt
    return c

def _banner(ws, title, subtitle, last_col):
    end = get_column_letter(last_col)
    ws.merge_cells(f"A1:{end}1")
    c = ws.cell(1, 1, f"  {title}")
    c.font      = _font(size=14, bold=True, color="FFFFFF")
    c.fill      = _fill(C_NAVY)
    c.alignment = _LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{end}2")
    c = ws.cell(2, 1, f"  {subtitle}")
    c.font      = _font(size=9, italic=True, color="BBBBBB")
    c.fill      = _fill(C_NAVY)
    c.alignment = _LEFT
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

# ── HCM Input sheet layout constants ─────────────────────────────────────────
# Row numbers in "HCM Input" sheet
R_GT_HDR  = 4
R_GT_DIRS = 5
R_GT_AM   = 6   # ← NEW input: AM green times
R_GT_PM   = 7   # ← NEW input: PM green times

R_PAR_HDR = 9
R_T       = 10  # ← NEW input
R_K       = 11  # ← NEW input
R_PF      = 12  # ← NEW input
R_I       = 13  # ← NEW input
R_L       = 14  # ← NEW input

R_REF_HDR = 16
R_REF_DIR = 17
R_AM_TOT  = 18  # linked from main sheet
R_PM_TOT  = 19  # linked from main sheet
R_LANES   = 20  # linked from main sheet
R_SAT     = 21  # linked from main sheet (V36)
R_CYCLE   = 22  # linked from main sheet (Z37)


def _build_input_sheet(wb):
    ws = wb.create_sheet("HCM Input")
    ws.sheet_properties.tabColor = "0D1B2A"

    # Column widths
    ws.column_dimensions["A"].width = 34
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["F"].width = 40

    _banner(ws,
        "HCM 7th Edition — Signal Analysis  |  Input Sheet",
        "Yellow cells = new inputs to fill in.  "
        "Blue cells = automatically linked from the main volume calculator sheet.",
        6)

    # ── GREEN TIMES ───────────────────────────────────────────────────────────
    _section_hdr(ws, R_GT_HDR, "NEW INPUT — EFFECTIVE GREEN TIMES  (seconds per approach)", 6)
    ws.row_dimensions[R_GT_DIRS].height = 18
    ws.cell(R_GT_DIRS, 1).value = ""
    for d in DIRS:
        _hdr_cell(ws, R_GT_DIRS, DIR_COL_N[d], DIR_EN[d], bg=C_DIRHEAD)
    ws.cell(R_GT_DIRS, 6, "← seconds of effective green per approach").font = \
        _font(size=9, italic=True, color=C_DKGREY)

    for row, label, default in [
        (R_GT_AM, "AM  effective green  (s)", 40),
        (R_GT_PM, "PM  effective green  (s)", 40),
    ]:
        ws.row_dimensions[row].height = 16
        ws.cell(row, 1, label).font = _font()
        for d in DIRS:
            _input_cell(ws, row, DIR_COL_N[d], default)
    ws.cell(R_GT_AM, 6,
        "Effective green = displayed green + yellow − start lost time"
    ).font = _font(size=9, italic=True, color=C_DKGREY)

    ws.row_dimensions[8].height = 6

    # ── SIGNAL PARAMETERS ─────────────────────────────────────────────────────
    _section_hdr(ws, R_PAR_HDR, "NEW INPUT — SIGNAL PARAMETERS", 6)

    params = [
        (R_T,  "Analysis period  T",           0.25, "hr",          "0.25 hr = 15-min peak period"),
        (R_K,  "Incremental delay factor  k",  0.5,  "—",           "0.50 = pre-timed  /  0.25 = actuated"),
        (R_PF, "Progression factor  PF",        1.0,  "—",           "1.0 = random arrivals  (HCM Table 19-9)"),
        (R_I,  "Upstream filter factor  I",     1.0,  "—",           "1.0 = isolated intersection"),
        (R_L,  "Car length  l",                 7,    "m",           "Used for Q95 in metres"),
    ]
    for row, label, val, unit, note in params:
        ws.row_dimensions[row].height = 16
        ws.cell(row, 1, label).font = _font()
        _input_cell(ws, row, 2, val)
        ws.cell(row, 3, unit).font = _font(size=9, color=C_DKGREY)
        ws.merge_cells(f"D{row}:F{row}")
        c = ws.cell(row, 4, note)
        c.font      = _font(size=9, italic=True, color=C_DKGREY)
        c.alignment = _LEFT

    ws.row_dimensions[15].height = 6

    # ── REFERENCES FROM MAIN SHEET ────────────────────────────────────────────
    _section_hdr(ws, R_REF_HDR,
        "AUTO-LINKED FROM MAIN SHEET  (do not edit — read-only)", 6,
        bg="1A3A6B")
    ws.row_dimensions[R_REF_DIR].height = 18
    ws.cell(R_REF_DIR, 1).value = ""
    for d in DIRS:
        _hdr_cell(ws, R_REF_DIR, DIR_COL_N[d], DIR_EN[d], bg=C_DIRHEAD)

    ref_rows = [
        (R_AM_TOT, "AM total volume  (veh/h)",  lambda d: _vol_ref(d, 4)),
        (R_PM_TOT, "PM total volume  (veh/h)",  lambda d: _vol_ref(d, 5)),
        (R_LANES,  "Total lanes",               lambda d: _lane_ref(d)),
    ]
    for row, label, formula_fn in ref_rows:
        ws.row_dimensions[row].height = 16
        ws.cell(row, 1, label).font = _font()
        for d in DIRS:
            _ref_cell(ws, row, DIR_COL_N[d], formula_fn(d))

    # Saturation flow (scalar — V36 on main sheet)
    ws.row_dimensions[R_SAT].height = 16
    ws.cell(R_SAT, 1, "Saturation flow  s").font = _font()
    _ref_cell(ws, R_SAT, 2, f"='{MAIN}'!V36", fmt="0")
    ws.cell(R_SAT, 3, "veh/h/lane").font = _font(size=9, color=C_DKGREY)
    ws.merge_cells(f"D{R_SAT}:F{R_SAT}")
    ws.cell(R_SAT, 4, f"Linked from main sheet cell V36").font = \
        _font(size=9, italic=True, color=C_DKGREY)

    # Cycle time (scalar — Z37 on main sheet)
    ws.row_dimensions[R_CYCLE].height = 16
    ws.cell(R_CYCLE, 1, "Cycle time  C").font = _font()
    _ref_cell(ws, R_CYCLE, 2, f"='{MAIN}'!Z37", fmt="0")
    ws.cell(R_CYCLE, 3, "s").font = _font(size=9, color=C_DKGREY)
    ws.merge_cells(f"D{R_CYCLE}:F{R_CYCLE}")
    ws.cell(R_CYCLE, 4, "Linked from main sheet cell Z37").font = \
        _font(size=9, italic=True, color=C_DKGREY)

    # ── LOS thresholds reference ──────────────────────────────────────────────
    ws.row_dimensions[24].height = 8
    _section_hdr(ws, 25, "LOS THRESHOLDS  (HCM 7th Ed. — Signalized Intersections)", 6)
    los_ref = [
        ("A", "≤ 10 s/veh",  "C6EFCE", "1E7A47", "Very low delay — excellent progression"),
        ("B", "10–20 s/veh", "DDEEFF", "1A5276", "Good progression or short cycle"),
        ("C", "20–35 s/veh", "FFF9C4", "7D6608", "Fair progression — higher delays"),
        ("D", "35–55 s/veh", "FFE0B2", "B15500", "Congestion influence noticeable"),
        ("E", "55–80 s/veh", "FFCDD2", "B71C1C", "High delay — poor progression or long cycle"),
        ("F", "> 80 s/veh",  "CC2200", "FFFFFF", "Unacceptable delay — over-capacity"),
    ]
    for i, (los, thr, bg, fg, desc) in enumerate(los_ref):
        row = 26 + i
        ws.row_dimensions[row].height = 14
        for col, val in [(1, los), (2, thr)]:
            c = ws.cell(row, col, val)
            c.font      = _font(size=9, bold=(col == 1), color=fg)
            c.fill      = _fill(bg)
            c.alignment = _CENTER
            c.border    = _THIN
        ws.merge_cells(f"C{row}:F{row}")
        c = ws.cell(row, 3, desc)
        c.font      = _font(size=9, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _LEFT
        c.border    = _THIN

    ws.freeze_panes = "B6"


# ── Data rows used by both result sheets ──────────────────────────────────────
# (direction, period, dir_col_letter, green_row, vol_row, lanes_row)
# All reference "HCM Input" sheet, not the main sheet directly.
DATA_ROWS = [
    ("N", "AM", "B", R_GT_AM, R_AM_TOT, R_LANES),
    ("N", "PM", "B", R_GT_PM, R_PM_TOT, R_LANES),
    ("S", "AM", "C", R_GT_AM, R_AM_TOT, R_LANES),
    ("S", "PM", "C", R_GT_PM, R_PM_TOT, R_LANES),
    ("E", "AM", "D", R_GT_AM, R_AM_TOT, R_LANES),
    ("E", "PM", "D", R_GT_PM, R_PM_TOT, R_LANES),
    ("W", "AM", "E", R_GT_AM, R_AM_TOT, R_LANES),
    ("W", "PM", "E", R_GT_PM, R_PM_TOT, R_LANES),
]

INP = "HCM Input"   # short alias for cross-sheet refs in formulas


def _common_cols(ws, r, d, period, dcol, green_row, vol_row, lanes_row):
    """Write cols A–H (Direction, Period, g, v, n, u, c, X) — shared by both result sheets."""
    bg = DIR_BG[d]
    for col, val, bold in [(1, DIR_EN[d], True), (2, period, False)]:
        c = ws.cell(r, col, val)
        c.font = _font(bold=bold)
        c.fill = _fill(bg)
        c.alignment = _CENTER
        c.border = _THIN

    # C: effective green g
    c = ws.cell(r, 3, f"='{INP}'!${dcol}${green_row}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # D: total approach volume v
    c = ws.cell(r, 4, f"='{INP}'!${dcol}${vol_row}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # E: total lanes n
    c = ws.cell(r, 5, f"='{INP}'!${dcol}${lanes_row}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # F: u = g/C
    _calc_cell(ws, r, 6,
        f"=IF('{INP}'!$B${R_CYCLE}=0,0,"
        f"C{r}/'{INP}'!$B${R_CYCLE})", fmt="0.000")

    # G: capacity c = s·n·u
    _calc_cell(ws, r, 7,
        f"='{INP}'!$B${R_SAT}*E{r}*F{r}", fmt="0")

    # H: X = v/c
    _calc_cell(ws, r, 8,
        f"=IF(G{r}=0,0,D{r}/G{r})", fmt="0.000")


def _build_delay_sheet(wb):
    ws = wb.create_sheet("HCM Delay")
    ws.sheet_properties.tabColor = "1B3A5C"
    NCOLS = 12

    col_widths = [11, 7, 10, 14, 9, 9, 13, 9, 10, 10, 13, 7]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _banner(ws,
        "HCM 7th Edition — Signal Delay & Level of Service",
        "d = d₁·PF + d₂     d₁ = 0.5·C·(1−u)²/(1−min(1,X)·u)     "
        "d₂ = 900T·[(X−1)+√((X−1)²+8kIX/(c·T))]     "
        "Edit green times and parameters in HCM Input sheet.",
        NCOLS)

    # Column headers
    hdrs = [
        "Direction", "Period", "Green g\n(s)", "Volume v\n(veh/h)",
        "Lanes n", "u = g/C", "Capacity c\n(veh/h)", "X = v/c",
        "d₁\n(s/veh)", "d₂\n(s/veh)", "Delay\n(s/veh)", "LOS",
    ]
    ws.row_dimensions[4].height = 34
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 4, col, h, wrap=True)

    for i, (d, period, dcol, green_row, vol_row, lanes_row) in enumerate(DATA_ROWS):
        r = 5 + i
        ws.row_dimensions[r].height = 18
        _common_cols(ws, r, d, period, dcol, green_row, vol_row, lanes_row)

        # I: d1 = 0.5·C·(1−u)² / (1−min(1,X)·u)
        _calc_cell(ws, r, 9,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"0.5*'{INP}'!$B${R_CYCLE}*(1-F{r})^2"
            f"/MAX(0.001,1-MIN(1,H{r})*F{r}))",
            fmt="0.0")

        # J: d2 = 900T·[(X−1)+√((X−1)²+8kIX/(c·T))]
        _calc_cell(ws, r, 10,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"900*'{INP}'!$B${R_T}*((H{r}-1)"
            f"+SQRT(MAX(0,(H{r}-1)^2"
            f"+8*'{INP}'!$B${R_K}*'{INP}'!$B${R_I}*H{r}"
            f"/(G{r}*'{INP}'!$B${R_T})))))",
            fmt="0.0")

        # K: Delay = d1·PF + d2
        _calc_cell(ws, r, 11,
            f"=I{r}*'{INP}'!$B${R_PF}+J{r}", fmt="0.0", bold=True)

        # L: LOS
        c = ws.cell(r, 12,
            f'=IF(K{r}<=10,"A",IF(K{r}<=20,"B",'
            f'IF(K{r}<=35,"C",IF(K{r}<=55,"D",'
            f'IF(K{r}<=80,"E","F")))))')
        c.font      = _font(bold=True)
        c.alignment = _CENTER
        c.border    = _THIN

    # Conditional formatting on LOS column (L5:L12)
    for los, (bg, fg) in LOS_STYLE.items():
        ws.conditional_formatting.add(
            "L5:L12",
            CellIsRule(
                operator="equal",
                formula=[f'"{los}"'],
                fill=_fill(bg),
                font=_font(bold=True, color=fg),
            )
        )

    # LOS legend
    ws.row_dimensions[14].height = 8
    ws.row_dimensions[15].height = 16
    ws.cell(15, 1, "LOS:").font = _font(bold=True)
    ws.cell(15, 1).alignment = _CENTER
    labels = {
        "A": "A  ≤ 10 s",  "B": "B  10–20 s", "C": "C  20–35 s",
        "D": "D  35–55 s", "E": "E  55–80 s",  "F": "F  > 80 s",
    }
    for i, (los, (bg, fg)) in enumerate(LOS_STYLE.items()):
        c = ws.cell(15, 2 + i, labels[los])
        c.font      = _font(size=9, bold=True, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _CENTER
        c.border    = _THIN

    ws.freeze_panes = "C5"


def _build_queue_sheet(wb):
    ws = wb.create_sheet("HCM 95th Queue")
    ws.sheet_properties.tabColor = "2C3E6B"
    NCOLS = 13

    col_widths = [11, 7, 10, 14, 9, 9, 13, 9, 11, 11, 11, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _banner(ws,
        "HCM 7th Edition — 95th Percentile Back-of-Queue",
        "Nq = Nq₁ + Nq₂     Q95 = Nq + 1.65·√Nq     "
        "Q95 (m) = Q95 · car length l     "
        "Edit green times and parameters in HCM Input sheet.",
        NCOLS)

    hdrs = [
        "Direction", "Period", "Green g\n(s)", "Volume v\n(veh/h)",
        "Lanes n", "u = g/C", "Capacity c\n(veh/h)", "X = v/c",
        "Nq₁\n(uniform)", "Nq₂\n(overflow)", "Nq avg\n(veh)", "Q95\n(veh)", "Q95\n(m)",
    ]
    ws.row_dimensions[4].height = 34
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 4, col, h, wrap=True)

    for i, (d, period, dcol, green_row, vol_row, lanes_row) in enumerate(DATA_ROWS):
        r = 5 + i
        ws.row_dimensions[r].height = 18
        _common_cols(ws, r, d, period, dcol, green_row, vol_row, lanes_row)

        # I: Nq1 = c·C/3600·(1−u)² / (1−min(1,X)·u)
        _calc_cell(ws, r, 9,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"G{r}*'{INP}'!$B${R_CYCLE}/3600"
            f"*(1-F{r})^2"
            f"/MAX(0.001,1-MIN(1,H{r})*F{r}))",
            fmt="0.0")

        # J: Nq2 = 900T·[(X−1)+√(...)]·c/3600
        _calc_cell(ws, r, 10,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"900*'{INP}'!$B${R_T}*((H{r}-1)"
            f"+SQRT(MAX(0,(H{r}-1)^2"
            f"+8*'{INP}'!$B${R_K}*'{INP}'!$B${R_I}*H{r}"
            f"/(G{r}*'{INP}'!$B${R_T}))))"
            f"*G{r}/3600)",
            fmt="0.0")

        # K: Nq = max(0, Nq1+Nq2)
        _calc_cell(ws, r, 11,
            f"=MAX(0,I{r}+J{r})", fmt="0.0")

        # L: Q95 in vehicles
        _calc_cell(ws, r, 12,
            f"=K{r}+1.65*SQRT(MAX(0,K{r}))", fmt="0.0", bold=True)

        # M: Q95 in metres
        _calc_cell(ws, r, 13,
            f"=L{r}*'{INP}'!$B${R_L}", fmt="0", bold=True)

    ws.freeze_panes = "C5"


# ── Main ──────────────────────────────────────────────────────────────────────
def add_hcm_sheets(src=SRC):
    if not os.path.exists(src):
        raise FileNotFoundError(f"Template not found: {src}")

    # Safety backup
    bak = src.replace(".xlsx", "_backup_before_hcm.xlsx")
    shutil.copy2(src, bak)
    print(f"Backup: {bak}")

    wb = xl.load_workbook(src)

    # Remove existing HCM sheets if re-running
    for name in ["HCM Input", "HCM Delay", "HCM 95th Queue"]:
        if name in wb.sheetnames:
            del wb[name]
            print(f"Removed existing sheet: {name}")

    _build_input_sheet(wb)
    _build_delay_sheet(wb)
    _build_queue_sheet(wb)

    wb.save(src)
    print(f"Saved: {src}")
    print("Sheets added: HCM Input, HCM Delay, HCM 95th Queue")


if __name__ == "__main__":
    add_hcm_sheets()
