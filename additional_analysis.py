"""
additional_analysis.py — HCM traffic-engineering analyses for JUNC.

Public API (imported by app.py and pipeline.py):

  make_queue_excel(car_length_dict, queue_params) → bytes
      Formatted .xlsx summarising Poisson queue-length results.

  compute_hcm_delay(state, green_times, params) → dict
      HCM 6th edition uniform + overflow signal delay and LOS per approach.

  compute_hcm_queue95(state, green_times, params) → dict
      HCM 95th-percentile back-of-queue per approach.

  make_hcm_excel(delay_results, queue95_results, green_times, params) → bytes
      Two-sheet .xlsx (Delay & LOS  +  95th Queue).
"""

import math
from io import BytesIO

DIRECTIONS = ["N", "S", "E", "W"]
_DIR_EN = {"N": "North", "S": "South", "E": "East",  "W": "West"}
_DIR_HE = {"N": "צפון",  "S": "דרום",  "E": "מזרח", "W": "מערב"}

_MOVEMENTS = ["R", "RT", "T", "TL", "L", "RTL", "RL"]
_MOVE_EN = {
    "R":   "Right",
    "RT":  "Right + Through",
    "T":   "Through",
    "TL":  "Through + Left",
    "L":   "Left",
    "RTL": "Right + Through + Left",
    "RL":  "Right + Left",
}
_MOVE_HE = {
    "R":   "ימינה",
    "RT":  "ימינה ישר",
    "T":   "ישר",
    "TL":  "ישר שמאלה",
    "L":   "שמאלה",
    "RTL": "הכול",
    "RL":  "ימינה ושמאלה",
}

# Queue-length movement names (matches Phaser.py ordering)
_QUEUE_NAMES = [
    "NR_length",  "NRT_length", "NT_length",  "NTL_length", "NL_length",  "NRTL_length", "NRL_length",
    "SR_length",  "SRT_length", "ST_length",  "STL_length", "SL_length",  "SRTL_length", "SRL_length",
    "ER_length",  "ERT_length", "ET_length",  "ETL_length", "EL_length",  "ERTL_length", "ERL_length",
    "WR_length",  "WRT_length", "WT_length",  "WTL_length", "WL_length",  "WRTL_length", "WRL_length",
]


def _los(delay: float) -> str:
    if delay <= 10:  return "A"
    if delay <= 20:  return "B"
    if delay <= 35:  return "C"
    if delay <= 55:  return "D"
    if delay <= 80:  return "E"
    return "F"


def _total_lanes(lane_dict: dict) -> int:
    """Total physical lanes for an approach (minimum 1)."""
    return max(1, sum(v for v in lane_dict.values() if v > 0))


# ── Poisson Queue Excel ────────────────────────────────────────────────────────

def make_queue_excel(car_length_dict: dict, queue_params: dict) -> bytes:
    """Return a formatted .xlsx file summarising Poisson queue-length results."""
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Queue Lengths"
    ws.sheet_view.rightToLeft = False

    title_font = Font(bold=True, size=14)
    hdr_font   = Font(bold=True, size=11)
    param_font = Font(italic=True, size=10, color="555555")
    zero_font  = Font(color="BBBBBB", size=10)
    data_font  = Font(size=10)
    bold_font  = Font(bold=True, size=10)

    dir_fill = {
        "N": PatternFill("solid", fgColor="DDEEFF"),
        "S": PatternFill("solid", fgColor="D6F5D6"),
        "E": PatternFill("solid", fgColor="FFF2CC"),
        "W": PatternFill("solid", fgColor="FFE0E0"),
    }
    hdr_fill   = PatternFill("solid", fgColor="333355")
    param_fill = PatternFill("solid", fgColor="F5F5F5")

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18

    row = 1
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row, 1, "Queue Length Analysis — JUNC")
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[row].height = 24
    row += 2

    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row, 1, "Calculation parameters").font = hdr_font
    row += 1

    param_rows = [
        ("Discard green time",         queue_params.get("discard_green_time", False)),
        ("Basic lost capacity (veh/h)", queue_params.get("basic_lost_capacity", 200)),
        ("Poisson percentile",          queue_params.get("poisson", 0.95)),
        ("Car length — l (m)",          queue_params.get("l", 7)),
        ("Peak hour factor — PHF",      queue_params.get("phf", 0.9)),
        ("Cycle time (s)",              queue_params.get("cycle_time", 120)),
    ]
    for label, val in param_rows:
        ws.cell(row, 1, label).font = param_font
        ws.cell(row, 1).fill = param_fill
        ws.cell(row, 2, str(val)).font = param_font
        ws.cell(row, 2).fill = param_fill
        row += 1
    row += 1

    headers = ["Direction", "Direction (HE)", "Movement", "Movement (HE)", "Queue Length (m)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = border
    ws.row_dimensions[row].height = 18
    row += 1

    for direction in ["N", "S", "E", "W"]:
        fill = dir_fill[direction]
        for movement in _MOVEMENTS:
            key    = f"{direction}{movement}_length"
            length = car_length_dict.get(key, 0)
            is_zero = (length == 0)
            font_to_use = zero_font if is_zero else data_font
            cells = [
                ws.cell(row, 1, _DIR_EN[direction]),
                ws.cell(row, 2, _DIR_HE[direction]),
                ws.cell(row, 3, _MOVE_EN[movement]),
                ws.cell(row, 4, _MOVE_HE[movement]),
                ws.cell(row, 5, length if not is_zero else "—"),
            ]
            for c in cells:
                c.fill = fill
                c.font = font_to_use
                c.border = border
                c.alignment = center
            if not is_zero:
                ws.cell(row, 5).font = bold_font
            row += 1

    hdr_row = row - 4 * len(_MOVEMENTS) - 1
    ws.auto_filter.ref = f"A{hdr_row}:E{row - 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── HCM Signal Delay & LOS ────────────────────────────────────────────────────

def compute_hcm_delay(state: dict, green_times: dict, params: dict) -> dict:
    """
    HCM 6th edition uniform + overflow signal delay and LOS per approach.

    Parameters
    ----------
    state : dict
        JUNC state dict (requires "volumes" and "lanes" sub-dicts).
    green_times : dict
        {"N": {"am": g, "pm": g}, "S": ..., "E": ..., "W": ...}
        Effective green time in seconds per approach per period.
    params : dict
        cycle_time (s), saturation_flow (veh/h/lane), T (analysis period, hr),
        k (incremental delay factor), PF (progression factor), I (upstream filter).

    Returns
    -------
    dict[direction][period] → {v, c, X, d1, d2, delay, LOS, g}
    """
    C  = params["cycle_time"]
    s  = params["saturation_flow"]
    T  = params["T"]
    k  = params["k"]
    PF = params["PF"]
    Iu = params["I"]

    results = {}
    for d in DIRECTIONS:
        results[d] = {}
        n_lanes = _total_lanes(state["lanes"][d])
        s_total = s * n_lanes

        for period, label in [("morning", "am"), ("evening", "pm")]:
            v = sum(state["volumes"][period][d].values())
            g = green_times[d][label]
            u = g / C if C > 0 else 0.0

            c = s_total * u
            if c == 0 or v == 0:
                results[d][period] = {
                    "v": v, "c": round(c), "X": 0.0,
                    "d1": 0.0, "d2": 0.0, "delay": 0.0, "LOS": "A", "g": g,
                }
                continue

            X  = v / c
            Xc = min(1.0, X)

            # Uniform delay (Webster / HCM d1)
            denom1 = 1.0 - Xc * u
            d1 = (0.5 * C * (1.0 - u) ** 2 / denom1) if denom1 > 0 else 0.0

            # Overflow delay (HCM d2)
            inner = (X - 1.0) ** 2 + 8.0 * k * Iu * X / (c * T)
            d2 = 900.0 * T * ((X - 1.0) + math.sqrt(max(0.0, inner)))

            delay = d1 * PF + d2

            results[d][period] = {
                "v":     v,
                "c":     round(c),
                "X":     round(X, 3),
                "d1":    round(d1, 1),
                "d2":    round(d2, 1),
                "delay": round(delay, 1),
                "LOS":   _los(delay),
                "g":     g,
            }
    return results


# ── HCM 95th-Percentile Queue ─────────────────────────────────────────────────

def compute_hcm_queue95(state: dict, green_times: dict, params: dict) -> dict:
    """
    HCM 95th-percentile back-of-queue per approach per period.

    Same parameters as compute_hcm_delay, plus params["l"] (car length in metres).

    Returns
    -------
    dict[direction][period] → {v, c, X, Nq, Q95_veh, Q95_m, g}
    """
    C  = params["cycle_time"]
    s  = params["saturation_flow"]
    T  = params["T"]
    k  = params["k"]
    Iu = params["I"]
    l  = params["l"]

    results = {}
    for d in DIRECTIONS:
        results[d] = {}
        n_lanes = _total_lanes(state["lanes"][d])
        s_total = s * n_lanes

        for period, label in [("morning", "am"), ("evening", "pm")]:
            v = sum(state["volumes"][period][d].values())
            g = green_times[d][label]
            u = g / C if C > 0 else 0.0

            c = s_total * u
            if c == 0 or v == 0:
                results[d][period] = {
                    "v": v, "c": round(c), "X": 0.0,
                    "Nq": 0.0, "Q95_veh": 0.0, "Q95_m": 0, "g": g,
                }
                continue

            X  = v / c
            Xc = min(1.0, X)

            # Average uniform queue (vehicles at end of red)
            denom1 = 1.0 - Xc * u
            Nq1 = (c * C / 3600.0 * (1.0 - u) ** 2 / denom1) if denom1 > 0 else 0.0

            # Average overflow queue
            inner = (X - 1.0) ** 2 + 8.0 * k * Iu * X / (c * T)
            Nq2   = 900.0 * T * ((X - 1.0) + math.sqrt(max(0.0, inner))) * c / 3600.0

            Nq      = max(0.0, Nq1 + Nq2)
            Q95_veh = Nq + 1.65 * math.sqrt(max(0.0, Nq))
            Q95_m   = Q95_veh * l

            results[d][period] = {
                "v":       v,
                "c":       round(c),
                "X":       round(X, 3),
                "Nq":      round(Nq, 1),
                "Q95_veh": round(Q95_veh, 1),
                "Q95_m":   round(Q95_m),
                "g":       g,
            }
    return results


# ── HCM Excel export ──────────────────────────────────────────────────────────

def make_hcm_excel(
    delay_results: dict,
    queue95_results: dict,
    green_times: dict,
    params: dict,
) -> bytes:
    """
    Build a two-sheet .xlsx:
      Sheet 1 — HCM Signal Delay & LOS
      Sheet 2 — HCM 95th Percentile Queue
    Returns raw bytes.
    """
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = xl.Workbook()

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    hdr_font   = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor="333355")
    title_font = Font(bold=True, size=13)
    param_font = Font(italic=True, size=9, color="555555")
    param_fill = PatternFill("solid", fgColor="F5F5F5")

    dir_fill = {
        "N": PatternFill("solid", fgColor="DDEEFF"),
        "S": PatternFill("solid", fgColor="D6F5D6"),
        "E": PatternFill("solid", fgColor="FFF2CC"),
        "W": PatternFill("solid", fgColor="FFE0E0"),
    }
    los_colors = {
        "A": "C6EFCE", "B": "FFEB9C", "C": "FFC7CE",
        "D": "FF9999", "E": "FF6666", "F": "AA0000",
    }

    def _write_title(ws, text, ncols):
        end_col = chr(64 + ncols)
        ws.merge_cells(f"A1:{end_col}1")
        c = ws.cell(1, 1, text)
        c.font = title_font
        c.alignment = center
        ws.row_dimensions[1].height = 22

    def _write_params(ws, start_row, include_pf=True, include_l=False):
        rows = [
            ("Cycle time (s)",               params.get("cycle_time", 120)),
            ("Saturation flow (veh/h/lane)", params.get("saturation_flow", 1800)),
            ("Analysis period T (hr)",       params.get("T", 0.25)),
            ("Incremental delay factor k",   params.get("k", 0.5)),
        ]
        if include_pf:
            rows.append(("Progression factor PF", params.get("PF", 1.0)))
        rows.append(("Upstream filter I", params.get("I", 1.0)))
        if include_l:
            rows.append(("Car length l (m)", params.get("l", 7)))
        r = start_row
        for label, val in rows:
            ws.cell(r, 1, label).font = param_font
            ws.cell(r, 1).fill = param_fill
            ws.cell(r, 2, str(val)).font = param_font
            ws.cell(r, 2).fill = param_fill
            r += 1
        return r + 1  # blank separator row

    # ── Sheet 1: HCM Delay & LOS ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "HCM Delay & LOS"
    for col, w in zip("ABCDEFGHIJ", [12, 8, 10, 13, 14, 10, 10, 10, 14, 7]):
        ws1.column_dimensions[col].width = w

    _write_title(ws1, "HCM Signal Delay & Level of Service — JUNC", 10)
    data_row = _write_params(ws1, 3, include_pf=True, include_l=False)

    hdrs1 = [
        "Direction", "Period", "Green (s)", "Volume (veh/h)",
        "Capacity (veh/h)", "X (v/c)", "d1 (s)", "d2 (s)", "Delay (s/veh)", "LOS",
    ]
    for col, h in enumerate(hdrs1, 1):
        c = ws1.cell(data_row, col, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    ws1.row_dimensions[data_row].height = 18
    data_row += 1

    for d in DIRECTIONS:
        fill = dir_fill[d]
        for period, label in [("morning", "AM"), ("evening", "PM")]:
            r   = delay_results[d][period]
            g   = green_times[d]["am" if period == "morning" else "pm"]
            los = r["LOS"]
            los_fill = PatternFill("solid", fgColor=los_colors[los])
            vals = [_DIR_EN[d], label, g, r["v"], r["c"],
                    r["X"], r["d1"], r["d2"], r["delay"], los]
            for col, val in enumerate(vals, 1):
                cell = ws1.cell(data_row, col, val)
                cell.border = border
                cell.alignment = center
                if col == 10:
                    cell.fill = los_fill
                    cell.font = Font(bold=True, size=10,
                                     color="FFFFFF" if los in ("E", "F") else "000000")
                else:
                    cell.fill = fill
                    cell.font = Font(bold=(col == 9), size=10)
            data_row += 1

    # ── Sheet 2: HCM 95th Queue ───────────────────────────────────────────────
    ws2 = wb.create_sheet("HCM 95th Queue")
    for col, w in zip("ABCDEFGHI", [12, 8, 10, 13, 14, 10, 14, 12, 10]):
        ws2.column_dimensions[col].width = w

    _write_title(ws2, "HCM 95th Percentile Back-of-Queue — JUNC", 9)
    data_row = _write_params(ws2, 3, include_pf=False, include_l=True)

    hdrs2 = [
        "Direction", "Period", "Green (s)", "Volume (veh/h)",
        "Capacity (veh/h)", "X (v/c)", "Avg Queue (veh)", "Q95 (veh)", "Q95 (m)",
    ]
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(data_row, col, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    ws2.row_dimensions[data_row].height = 18
    data_row += 1

    for d in DIRECTIONS:
        fill = dir_fill[d]
        for period, label in [("morning", "AM"), ("evening", "PM")]:
            r    = queue95_results[d][period]
            g    = green_times[d]["am" if period == "morning" else "pm"]
            vals = [_DIR_EN[d], label, g, r["v"], r["c"],
                    r["X"], r["Nq"], r["Q95_veh"], r["Q95_m"]]
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(data_row, col, val)
                cell.border = border
                cell.alignment = center
                cell.fill = fill
                cell.font = Font(bold=(col == 9), size=10)
            data_row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
