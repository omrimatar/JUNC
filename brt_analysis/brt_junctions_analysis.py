# -*- coding: utf-8 -*-
"""
ניתוח צמתים מרומזרים על צירי ה-BRT (קו כחול / קו חום שלב א') ובמרחק עד 200 מ' מהם.

קלט : שכבת צמתים מרומזרים (GeoJSON, EPSG:2039 – רשת ישראל החדשה).
       כל צומת מיוצג בעיגול ברדיוס 20 מ' סביב הנקודה (X, Y) של הצומת.
פלט : 1. brt_junctions.geojson          – נקודות הרמזורים שבמפגש עם הקו הכחול / החום (שלב א').
       2. junctions_within_200m.geojson  – צמתים מרומזרים נוספים במרחק אווירי <= 200 מ' מצומת BRT.
       3. brt_junctions_analysis.xlsx    – טבלאות מסכמות.

הגדרות (בהתאם לסיכום עם המזמין):
  * "צומת שהקו עובר בו" = צומת שבשכבה מסומן blue_line='כן' או brown_line='כן'.
  * קו חום – שלב א' בלבד (מערב ראשל"צ – מסוף צריפין / מרכז רפואי שמיר) => צמתים בראשון לציון בלבד.
    צמתי החום המזרחי (באר יעקב – נעמי שמר, רמלה, לוד) מוחרגים ומדווחים בנפרד.
  * מרחק = מרחק אווירי (אוקלידי) בין מרכזי הצמתים, במטרים, במערכת EPSG:2039.
"""
import json, math, sys, os
from collections import OrderedDict

SRC = sys.argv[1] if len(sys.argv) > 1 else "ramzor.geojson"
OUT_DIR = sys.argv[2] if len(sys.argv) > 2 else "."
RADIUS_M = 200.0

os.makedirs(OUT_DIR, exist_ok=True)
data = json.load(open(SRC, encoding="utf-8"))
feats = data["features"]
CRS = data.get("crs", {"type": "name", "properties": {"name": "urn:ogc:def:crs:EPSG::2039"}})


def is_yes(v):
    return v is not None and "כן" in str(v)


def clean(v):
    """השדות בשכבה מגיעים עם פסיקים מיותרים (תוצר מיזוג שכבות) – ניקוי לתצוגה."""
    if v is None:
        return ""
    s = str(v)
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return " ".join(OrderedDict.fromkeys(parts))


def junction_name(p):
    return clean(p.get("name")) or clean(p.get("JUNCTION")) or clean(p.get("STREETROAD"))


def streets(p):
    vals = [clean(p.get(k)) for k in ("main street", "stree_1", "stree_2", "street_3", "street_4")]
    vals = [v for v in vals if v]
    return " / ".join(OrderedDict.fromkeys(vals)) if vals else clean(p.get("STREETROAD"))


def city(p):
    return clean(p.get("CITY"))


# ---------- 1. סיווג צמתי BRT ----------
brt, brown_east_excluded = [], []
for f in feats:
    p = f["properties"]
    blue = is_yes(p.get("blue_line"))
    brown = is_yes(p.get("brown_line"))
    if not (blue or brown):
        continue
    # קו חום שלב א' = ראשון לציון בלבד. צומת 2463 (הרצל/רוטשילד) חסר עיר בשכבה אך נמצא בראשל"צ.
    brown_phase_a = brown and (("ראשון" in city(p)) or p["fid"] == 2463)
    if brown and not brown_phase_a and not blue:
        brown_east_excluded.append(f)
        continue
    if blue and brown_phase_a:
        line = "כחול + חום"
    elif blue:
        line = "כחול"
    else:
        line = "חום (שלב א')"
    brt.append((f, line))

brt_ids = {f["properties"]["fid"] for f, _ in brt}


def xy(f):
    return f["properties"]["X"], f["properties"]["Y"]


def dist(a, b):
    ax, ay = xy(a)
    bx, by = xy(b)
    return math.hypot(ax - bx, ay - by)


# ---------- 2. צמתים עד 200 מ' ----------
neighbors = {}  # fid -> dict
pairs = []
for f in feats:
    p = f["properties"]
    if p["fid"] in brt_ids:
        continue
    hits = []
    for bf, line in brt:
        d = dist(f, bf)
        if d <= RADIUS_M:
            hits.append((d, bf, line))
    if hits:
        hits.sort(key=lambda h: h[0])
        neighbors[p["fid"]] = {"feat": f, "hits": hits}
        for d, bf, line in hits:
            pairs.append((bf, line, f, d))

# מרחק לצומת ה-BRT הקרוב ביותר גם עבור צמתי ה-BRT עצמם (מידע עזר)
brt_nearest = {}
for bf, _ in brt:
    best = None
    for of, _ in brt:
        if of is bf:
            continue
        d = dist(bf, of)
        if best is None or d < best[0]:
            best = (d, of)
    brt_nearest[bf["properties"]["fid"]] = best


def point_feature(f, extra):
    p = dict(f["properties"])
    p.update(extra)
    return {"type": "Feature", "properties": p,
            "geometry": {"type": "Point", "coordinates": [p["X"], p["Y"]]}}


brt_out = []
for bf, line in brt:
    fid = bf["properties"]["fid"]
    nb_list = sorted([(d, nf) for b2, _, nf, d in pairs if b2 is bf], key=lambda t: t[0])
    nearest = brt_nearest[fid]
    brt_out.append(point_feature(bf, {
        "brt_line": line,
        "junction_name": junction_name(bf["properties"]),
        "streets": streets(bf["properties"]),
        "city_clean": city(bf["properties"]) or "ראשון לציון",
        "status_clean": clean(bf["properties"].get("Status")),
        "n_junctions_within_200m": len(nb_list),
        "junctions_within_200m_fids": ", ".join(str(nf["properties"]["fid"]) for _, nf in nb_list),
        "nearest_brt_junction_fid": nearest[1]["properties"]["fid"] if nearest else None,
        "nearest_brt_junction_dist_m": round(nearest[0], 1) if nearest else None,
    }))

nb_out = []
for fid, rec in sorted(neighbors.items()):
    f, hits = rec["feat"], rec["hits"]
    d0, bf0, line0 = hits[0]
    nb_out.append(point_feature(f, {
        "junction_name": junction_name(f["properties"]),
        "streets": streets(f["properties"]),
        "city_clean": city(f["properties"]),
        "status_clean": clean(f["properties"].get("Status")),
        "nearest_brt_fid": bf0["properties"]["fid"],
        "nearest_brt_name": junction_name(bf0["properties"]),
        "nearest_brt_line": line0,
        "dist_to_nearest_brt_m": round(d0, 1),
        "n_brt_within_200m": len(hits),
        "brt_fids_within_200m": ", ".join(str(h[1]["properties"]["fid"]) for h in hits),
    }))


def write_geojson(path, features, name):
    json.dump({"type": "FeatureCollection", "name": name, "crs": CRS, "features": features},
              open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)


write_geojson(os.path.join(OUT_DIR, "brt_junctions.geojson"), brt_out, "brt_junctions")
write_geojson(os.path.join(OUT_DIR, "junctions_within_200m.geojson"), nb_out, "junctions_within_200m")

# ---------- 3. Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
LINE_FILL = {"כחול": "DDEBF7", "חום (שלב א')": "F8E5D0", "כחול + חום": "E2D6EF"}


def add_sheet(title, headers, rows, widths=None, fill_col=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font, c.border = BODY_FONT, BORDER
        if fill_col is not None:
            col = LINE_FILL.get(row[fill_col].value)
            if col:
                for c in row:
                    c.fill = PatternFill("solid", fgColor=col)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = (widths[i - 1] if widths else 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


# גיליון 1 – צמתי BRT
h1 = ["fid", "קו", "עיר", "שם צומת", "רחובות", "סטטוס רמזור", "X (ITM)", "Y (ITM)",
      "מס' צמתים מרומזרים עד 200 מ'", "fid צמתים עד 200 מ'", "צומת BRT הקרוב ביותר (fid)", "מרחק לצומת BRT הקרוב (מ')"]
rows1 = []
order = {"כחול": 0, "כחול + חום": 1, "חום (שלב א')": 2}
for f in sorted(brt_out, key=lambda f: (order[f["properties"]["brt_line"]], -f["properties"]["Y"])):
    p = f["properties"]
    rows1.append([p["fid"], p["brt_line"], p["city_clean"], p["junction_name"], p["streets"], p["status_clean"],
                  round(p["X"], 1), round(p["Y"], 1), p["n_junctions_within_200m"], p["junctions_within_200m_fids"],
                  p["nearest_brt_junction_fid"], p["nearest_brt_junction_dist_m"]])
add_sheet("צמתי BRT", h1, rows1, [7, 13, 14, 34, 44, 12, 11, 11, 12, 22, 12, 12], fill_col=1)

# גיליון 2 – צמתים עד 200 מ'
h2 = ["fid", "עיר", "שם צומת", "רחובות", "סטטוס רמזור", "X (ITM)", "Y (ITM)",
      "צומת BRT הקרוב (fid)", "צומת BRT הקרוב (שם)", "קו", "מרחק אווירי (מ')", "מס' צמתי BRT עד 200 מ'", "fid צמתי BRT עד 200 מ'"]
rows2 = []
for f in sorted(nb_out, key=lambda f: f["properties"]["dist_to_nearest_brt_m"]):
    p = f["properties"]
    rows2.append([p["fid"], p["city_clean"], p["junction_name"], p["streets"], p["status_clean"], round(p["X"], 1), round(p["Y"], 1),
                  p["nearest_brt_fid"], p["nearest_brt_name"], p["nearest_brt_line"], p["dist_to_nearest_brt_m"],
                  p["n_brt_within_200m"], p["brt_fids_within_200m"]])
add_sheet("צמתים עד 200 מ'", h2, rows2, [7, 14, 34, 44, 12, 11, 11, 11, 30, 13, 11, 10, 16], fill_col=9)

# גיליון 3 – זוגות (BRT ↔ צומת סמוך)
h3 = ["fid צומת BRT", "קו", "שם צומת BRT", "fid צומת סמוך", "שם צומת סמוך", "עיר", "מרחק אווירי (מ')"]
rows3 = [[bf["properties"]["fid"], line, junction_name(bf["properties"]), nf["properties"]["fid"],
          junction_name(nf["properties"]), city(nf["properties"]), round(d, 1)]
         for bf, line, nf, d in sorted(pairs, key=lambda t: (order[t[1]], t[0]["properties"]["fid"], t[3]))]
add_sheet("זוגות BRT-סמוך", h3, rows3, [11, 13, 34, 11, 34, 14, 12], fill_col=1)

# גיליון 4 – חום מזרחי (מוחרג)
h4 = ["fid", "עיר", "שם צומת", "רחובות", "סטטוס רמזור", "X (ITM)", "Y (ITM)", "הערה"]
rows4 = [[f["properties"]["fid"], city(f["properties"]), junction_name(f["properties"]), streets(f["properties"]),
          clean(f["properties"].get("Status")), round(f["properties"]["X"], 1), round(f["properties"]["Y"], 1),
          "מסומן brown_line='כן' בשכבה – קו חום מזרחי, לא חלק משלב א' (לא מתוקצב במצגת)"]
         for f in sorted(brown_east_excluded, key=lambda f: f["properties"]["X"])]
add_sheet("חום מזרחי - מוחרג", h4, rows4, [7, 12, 34, 44, 12, 11, 11, 60])

# גיליון סיכום (ראשון)
ws = wb["Sheet"]
ws.title = "סיכום"
wb.move_sheet(ws, offset=-4)
ws.sheet_view.rightToLeft = True
n_blue = sum(1 for f in brt_out if f["properties"]["brt_line"] == "כחול")
n_brown = sum(1 for f in brt_out if f["properties"]["brt_line"] == "חום (שלב א')")
n_both = sum(1 for f in brt_out if f["properties"]["brt_line"] == "כחול + חום")
summary = [
    ["ניתוח צמתים מרומזרים – קו כחול וקו חום (BRT)", ""],
    ["", ""],
    ["מקור נתונים", "שכבת צמתים מרומזרים (ramzor.geojson, EPSG:2039) + תכנית תפעולית קו כחול וחום, ועדת מכרזים 24.6.2026"],
    ["הגדרת 'צומת שהקו עובר בו'", "צומת המסומן בשכבה blue_line='כן' או brown_line='כן'"],
    ["קו חום", "שלב א' בלבד (מערב ראשל\"צ – מסוף צריפין / מרכז רפואי שמיר). צמתי החום המזרחי (באר יעקב, רמלה, לוד) מוחרגים – ראה גיליון נפרד"],
    ["מדידת מרחק", "מרחק אווירי בין מרכזי הצמתים (נקודות X,Y), במטרים, ברשת ישראל. סף: עד 200 מ' כולל"],
    ["", ""],
    ["צמתי BRT – סה\"כ", len(rows1)],
    ["מתוכם קו כחול", n_blue],
    ["מתוכם קו חום (שלב א')", n_brown],
    ["מתוכם משותפים לשני הקווים", n_both],
    ["צמתים מרומזרים נוספים עד 200 מ' מצומת BRT", len(rows2)],
    ["צמתי חום מזרחי שהוחרגו", len(rows4)],
    ["", ""],
    ["קבצי GeoJSON", "brt_junctions.geojson ; junctions_within_200m.geojson (נקודות, EPSG:2039, כל תכונות המקור + שדות ניתוח)"],
]
for r in summary:
    ws.append(r)
for row in ws.iter_rows():
    for c in row:
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws["A1"].font = Font(name="Arial", bold=True, size=13)
for r in range(3, 16):
    ws.cell(r, 1).font = Font(name="Arial", bold=True, size=10)
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 110

wb.save(os.path.join(OUT_DIR, "brt_junctions_analysis.xlsx"))

print(f"BRT junctions: {len(brt_out)} (blue {n_blue}, brown A {n_brown}, both {n_both})")
print(f"Brown-east excluded: {len(brown_east_excluded)}")
print(f"Neighbors within {RADIUS_M:.0f} m: {len(nb_out)} ; pairs: {len(pairs)}")
