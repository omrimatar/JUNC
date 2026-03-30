"""
build_hcm7_tool.py
Run once to produce HCM7_Analysis_Tool.xlsx in the project folder.
"""

from io import BytesIO
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY    = "0D1B2A"
C_STEEL   = "1B3A5C"
C_DIRHEAD = "2C3E6B"
C_LIGHT   = "E3F2FD"
C_INPUT   = "FFFCE8"
C_CALC    = "EBF5FB"
C_GREY    = "F5F5F5"
C_DKGREY  = "666666"

LOS_STYLE = {
    "A": ("C6EFCE", "1E7A47"),
    "B": ("DDEEFF", "1A5276"),
    "C": ("FFF9C4", "7D6608"),
    "D": ("FFE0B2", "B15500"),
    "E": ("FFCDD2", "B71C1C"),
    "F": ("CC2200", "FFFFFF"),
}

DIR_BG = {"N": "E8F4FD", "S": "E8F5E9", "E": "FFF8E1", "W": "FCE4EC"}

DIRECTIONS = ["N", "S", "E", "W"]
DIR_EN     = {"N": "North", "S": "South", "E": "East", "W": "West"}
DIR_COL    = {"N": "B", "S": "C", "E": "D", "W": "E"}
DIR_COL_N  = {"N": 2,   "S": 3,   "E": 4,   "W": 5}
LANE_TYPES = ["R", "RT", "T", "TL", "L", "RTL", "RL"]

# ── Style helpers ─────────────────────────────────────────────────────────────
def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _border(color="CCCCCC"):
    s = _side(color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, italic=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_WRAP_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN   = _border()

def _hdr_cell(ws, row, col, text, bg=C_NAVY, size=10, wrap=False):
    c = ws.cell(row, col, text)
    c.font      = _font(size=size, bold=True, color="FFFFFF")
    c.fill      = _fill(bg)
    c.alignment = _WRAP_C if wrap else _CENTER
    c.border    = _border(color="33557A")
    return c

def _section_hdr(ws, row, text, last_col, bg=C_STEEL):
    ws.merge_cells(f"A{row}:{get_column_letter(last_col)}{row}")
    c = ws.cell(row, 1, f"  {text}")
    c.font      = _font(size=10, bold=True, color="FFFFFF")
    c.fill      = _fill(bg)
    c.alignment = _LEFT
    ws.row_dimensions[row].height = 17

def _input_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row, col, value)
    c.font      = _font(bold=True, color="1A3A6B")
    c.fill      = _fill(C_INPUT)
    c.alignment = _CENTER
    c.border    = _THIN
    if fmt:
        c.number_format = fmt
    return c

def _calc_cell(ws, row, col, formula, fmt=None, bold=False):
    c = ws.cell(row, col, formula)
    c.font      = _font(bold=bold)
    c.fill      = _fill(C_CALC)
    c.alignment = _CENTER
    c.border    = _THIN
    if fmt:
        c.number_format = fmt
    return c


# ── Input sheet row constants ──────────────────────────────────────────────────
R_TITLE     = 1
R_PROJ      = 2
R_VOL_HDR   = 4
R_VOL_DIRS  = 5
R_AM_R      = 6
R_AM_T      = 7
R_AM_L      = 8
R_AM_TOT    = 9
R_PM_R      = 11
R_PM_T      = 12
R_PM_L      = 13
R_PM_TOT    = 14
R_LANE_HDR  = 16
R_LANE_DIRS = 17
R_LANE_R    = 18
R_LANE_RT   = 19
R_LANE_T    = 20
R_LANE_TL   = 21
R_LANE_L    = 22
R_LANE_RTL  = 23
R_LANE_RL   = 24
R_LANE_TOT  = 25
R_GT_HDR    = 27
R_GT_DIRS   = 28
R_GT_AM     = 29
R_GT_PM     = 30
R_PAR_HDR   = 32
R_C         = 33   # cycle time
R_S         = 34   # saturation flow
R_T         = 35   # analysis period
R_K         = 36   # k factor
R_PF        = 37   # progression factor PF
R_I         = 38   # upstream filter I
R_L         = 39   # car length l


def _build_input_sheet(wb):
    ws = wb.active
    ws.title = "Input"

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["F"].width = 38

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, "  HCM 7th Edition — Signal Analysis Tool")
    c.font = _font(size=15, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY)
    c.alignment = _LEFT
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, "  Yellow cells = inputs you can edit.  Blue cells = calculated automatically.")
    c.font = _font(size=9, italic=True, color="AAAAAA")
    c.fill = _fill(C_NAVY)
    c.alignment = _LEFT
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 6

    # ── VOLUMES ───────────────────────────────────────────────────────────────
    _section_hdr(ws, R_VOL_HDR, "TRAFFIC VOLUMES  (vehicles per hour)", 6)
    ws.row_dimensions[R_VOL_DIRS].height = 18
    ws.cell(R_VOL_DIRS, 1).value = ""
    for d in DIRECTIONS:
        _hdr_cell(ws, R_VOL_DIRS, DIR_COL_N[d], DIR_EN[d], bg=C_DIRHEAD)
    c = ws.cell(R_VOL_DIRS, 6, "")
    c.fill = _fill(C_DIRHEAD)

    vol_rows = [
        (R_AM_R,  "AM  Right  (R)",    "morning", "R"),
        (R_AM_T,  "AM  Through  (T)",  "morning", "T"),
        (R_AM_L,  "AM  Left  (L)",     "morning", "L"),
        (R_PM_R,  "PM  Right  (R)",    "evening", "R"),
        (R_PM_T,  "PM  Through  (T)",  "evening", "T"),
        (R_PM_L,  "PM  Left  (L)",     "evening", "L"),
    ]
    for row, label, _, _ in vol_rows:
        ws.row_dimensions[row].height = 15
        ws.cell(row, 1, label).font = _font()
        for d in DIRECTIONS:
            _input_cell(ws, row, DIR_COL_N[d], 0)

    for tot_row, r_row, t_row, l_row, label in [
        (R_AM_TOT, R_AM_R, R_AM_T, R_AM_L, "AM  Total"),
        (R_PM_TOT, R_PM_R, R_PM_T, R_PM_L, "PM  Total"),
    ]:
        ws.row_dimensions[tot_row].height = 15
        c = ws.cell(tot_row, 1, label)
        c.font = _font(bold=True)
        for d in DIRECTIONS:
            lc = DIR_COL[d]
            _calc_cell(ws, tot_row, DIR_COL_N[d],
                       f"={lc}{r_row}+{lc}{t_row}+{lc}{l_row}", fmt="0")
        c = ws.cell(tot_row, 6, "← auto-sum of R + T + L")
        c.font = _font(size=9, italic=True, color=C_DKGREY)

    ws.row_dimensions[10].height = 6
    ws.row_dimensions[15].height = 6

    # ── LANES ─────────────────────────────────────────────────────────────────
    _section_hdr(ws, R_LANE_HDR, "LANE CONFIGURATION  (number of lanes per type)", 6)
    ws.row_dimensions[R_LANE_DIRS].height = 18
    for d in DIRECTIONS:
        _hdr_cell(ws, R_LANE_DIRS, DIR_COL_N[d], DIR_EN[d], bg=C_DIRHEAD)

    lane_labels = [
        (R_LANE_R,   "R   — Right only"),
        (R_LANE_RT,  "RT  — Right + Through"),
        (R_LANE_T,   "T   — Through only"),
        (R_LANE_TL,  "TL  — Through + Left"),
        (R_LANE_L,   "L   — Left only"),
        (R_LANE_RTL, "RTL — Right + Through + Left"),
        (R_LANE_RL,  "RL  — Right + Left"),
    ]
    for row, label in lane_labels:
        ws.row_dimensions[row].height = 15
        ws.cell(row, 1, label).font = _font()
        for d in DIRECTIONS:
            _input_cell(ws, row, DIR_COL_N[d], 0)

    ws.row_dimensions[R_LANE_TOT].height = 16
    c = ws.cell(R_LANE_TOT, 1, "Total Lanes")
    c.font = _font(bold=True)
    for d in DIRECTIONS:
        lc = DIR_COL[d]
        _calc_cell(ws, R_LANE_TOT, DIR_COL_N[d],
                   f"=SUM({lc}{R_LANE_R}:{lc}{R_LANE_RL})", fmt="0")
    ws.cell(R_LANE_TOT, 6, "← total lanes per approach").font = _font(size=9, italic=True, color=C_DKGREY)

    ws.row_dimensions[26].height = 6

    # ── GREEN TIMES ───────────────────────────────────────────────────────────
    _section_hdr(ws, R_GT_HDR, "EFFECTIVE GREEN TIMES  (seconds per approach)", 6)
    ws.row_dimensions[R_GT_DIRS].height = 18
    for d in DIRECTIONS:
        _hdr_cell(ws, R_GT_DIRS, DIR_COL_N[d], DIR_EN[d], bg=C_DIRHEAD)

    for row, label in [(R_GT_AM, "AM  green  (s)"), (R_GT_PM, "PM  green  (s)")]:
        ws.row_dimensions[row].height = 15
        ws.cell(row, 1, label).font = _font()
        for d in DIRECTIONS:
            _input_cell(ws, row, DIR_COL_N[d], 40)
    ws.cell(R_GT_AM, 6,
            "Effective green = displayed green + yellow − start lost time"
            ).font = _font(size=9, italic=True, color=C_DKGREY)

    ws.row_dimensions[31].height = 6

    # ── SIGNAL PARAMETERS ─────────────────────────────────────────────────────
    _section_hdr(ws, R_PAR_HDR, "SIGNAL PARAMETERS", 6)

    param_defs = [
        (R_C,  "Cycle time  C",                    120,  "s",           "Total signal cycle duration"),
        (R_S,  "Saturation flow  s",               1800, "veh/h/lane",  "Per lane; typical range 1,600–1,900"),
        (R_T,  "Analysis period  T",               0.25, "hr",          "0.25 hr = 15-min peak period"),
        (R_K,  "Incremental delay factor  k",      0.5,  "—",           "0.50 = pre-timed  /  0.25 = actuated"),
        (R_PF, "Progression factor  PF",           1.0,  "—",           "1.0 = random arrivals  (HCM Table 19-9)"),
        (R_I,  "Upstream filter factor  I",        1.0,  "—",           "1.0 = isolated intersection"),
        (R_L,  "Car length  l",                    7,    "m",           "Used for Q95 in metres (col M of 95th Queue sheet)"),
    ]
    for row, label, val, unit, note in param_defs:
        ws.row_dimensions[row].height = 16
        ws.cell(row, 1, label).font = _font()
        _input_cell(ws, row, 2, val)
        ws.cell(row, 3, unit).font = _font(size=9, color=C_DKGREY)
        ws.merge_cells(f"D{row}:F{row}")
        c = ws.cell(row, 4, note)
        c.font = _font(size=9, italic=True, color=C_DKGREY)
        c.alignment = _LEFT

    ws.row_dimensions[40].height = 8

    # ── LOS THRESHOLDS reference ──────────────────────────────────────────────
    _section_hdr(ws, 41, "LOS THRESHOLDS  (HCM 7th Ed. — Signalized Intersections)", 6)
    los_ref = [
        ("A", "≤ 10 s/veh",  "C6EFCE", "1E7A47", "Very low delay — excellent progression"),
        ("B", "10–20 s/veh", "DDEEFF", "1A5276", "Good progression or short cycle"),
        ("C", "20–35 s/veh", "FFF9C4", "7D6608", "Fair progression — higher delays"),
        ("D", "35–55 s/veh", "FFE0B2", "B15500", "Congestion influence noticeable"),
        ("E", "55–80 s/veh", "FFCDD2", "B71C1C", "High delay — poor progression or long cycle"),
        ("F", "> 80 s/veh",  "CC2200", "FFFFFF", "Unacceptable delay — over-capacity"),
    ]
    for i, (los, thr, bg, fg, desc) in enumerate(los_ref):
        row = 42 + i
        ws.row_dimensions[row].height = 14
        for col, val in [(1, los), (2, thr)]:
            c = ws.cell(row, col, val)
            c.font = _font(size=9, bold=(col == 1), color=fg)
            c.fill = _fill(bg)
            c.alignment = _CENTER
            c.border = _THIN
        ws.merge_cells(f"C{row}:F{row}")
        c = ws.cell(row, 3, desc)
        c.font = _font(size=9, color=fg)
        c.fill = _fill(bg)
        c.alignment = _LEFT
        c.border = _THIN

    ws.freeze_panes = "B6"


# ── Data row map (shared by both result sheets) ───────────────────────────────
DATA_ROWS = [
    ("N", "AM", "B", R_GT_AM, R_AM_TOT),
    ("N", "PM", "B", R_GT_PM, R_PM_TOT),
    ("S", "AM", "C", R_GT_AM, R_AM_TOT),
    ("S", "PM", "C", R_GT_PM, R_PM_TOT),
    ("E", "AM", "D", R_GT_AM, R_AM_TOT),
    ("E", "PM", "D", R_GT_PM, R_PM_TOT),
    ("W", "AM", "E", R_GT_AM, R_AM_TOT),
    ("W", "PM", "E", R_GT_PM, R_PM_TOT),
]


def _common_cols(ws, r, d, period, dcol, green_row, vol_row):
    """Write cols A–H (shared between Delay and Queue sheets)."""
    bg = DIR_BG[d]
    for col, val, bold in [(1, DIR_EN[d], True), (2, period, False)]:
        c = ws.cell(r, col, val)
        c.font = _font(bold=bold)
        c.fill = _fill(bg)
        c.alignment = _CENTER
        c.border = _THIN

    # C: green g
    c = ws.cell(r, 3, f"=Input!${dcol}${green_row}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # D: volume
    c = ws.cell(r, 4, f"=Input!${dcol}${vol_row}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # E: total lanes
    c = ws.cell(r, 5, f"=Input!${dcol}${R_LANE_TOT}")
    c.font = _font(); c.fill = _fill(bg); c.alignment = _CENTER
    c.border = _THIN; c.number_format = "0"

    # F: u = g/C
    _calc_cell(ws, r, 6,
               f"=IF(Input!$B${R_C}=0,0,C{r}/Input!$B${R_C})", fmt="0.000")

    # G: capacity c = s * n * u
    _calc_cell(ws, r, 7,
               f"=Input!$B${R_S}*E{r}*F{r}", fmt="0")

    # H: X = v/c
    _calc_cell(ws, r, 8,
               f"=IF(G{r}=0,0,D{r}/G{r})", fmt="0.000")


def _sheet_banner(ws, title, subtitle, ncols):
    end = get_column_letter(ncols)
    ws.merge_cells(f"A1:{end}1")
    c = ws.cell(1, 1, f"  {title}")
    c.font = _font(size=14, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{end}2")
    c = ws.cell(2, 1, f"  {subtitle}")
    c.font = _font(size=9, italic=True, color="AAAAAA")
    c.fill = _fill(C_NAVY); c.alignment = _LEFT
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6


def _build_delay_sheet(wb):
    ws = wb.create_sheet("Delay & LOS")
    NCOLS = 12

    col_widths = [11, 7, 10, 14, 9, 9, 13, 9, 10, 10, 13, 7]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _sheet_banner(ws,
        "HCM 7th Edition — Signal Delay & Level of Service",
        "d = d₁·PF + d₂     d₁ = uniform delay (Webster)     d₂ = overflow delay"
        "     Edit parameters in the Input sheet",
        NCOLS)

    # Column headers
    hdrs = [
        "Direction", "Period", "Green g\n(s)", "Volume v\n(veh/h)",
        "Lanes n", "u = g/C", "Capacity c\n(veh/h)", "X = v/c",
        "d₁\n(s/veh)", "d₂\n(s/veh)", "Delay\n(s/veh)", "LOS",
    ]
    ws.row_dimensions[4].height = 32
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 4, col, h, wrap=True)

    # Data rows
    for i, (d, period, dcol, green_row, vol_row) in enumerate(DATA_ROWS):
        r = 5 + i
        ws.row_dimensions[r].height = 18
        _common_cols(ws, r, d, period, dcol, green_row, vol_row)

        # I: d1
        _calc_cell(ws, r, 9,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"0.5*Input!$B${R_C}*(1-F{r})^2"
            f"/MAX(0.001,1-MIN(1,H{r})*F{r}))",
            fmt="0.0")

        # J: d2
        _calc_cell(ws, r, 10,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"900*Input!$B${R_T}*((H{r}-1)"
            f"+SQRT(MAX(0,(H{r}-1)^2"
            f"+8*Input!$B${R_K}*Input!$B${R_I}*H{r}"
            f"/(G{r}*Input!$B${R_T})))))",
            fmt="0.0")

        # K: Delay = d1*PF + d2
        _calc_cell(ws, r, 11, f"=I{r}*Input!$B${R_PF}+J{r}", fmt="0.0", bold=True)

        # L: LOS
        c = ws.cell(r, 12,
            f'=IF(K{r}<=10,"A",IF(K{r}<=20,"B",'
            f'IF(K{r}<=35,"C",IF(K{r}<=55,"D",'
            f'IF(K{r}<=80,"E","F")))))')
        c.font = _font(bold=True)
        c.alignment = _CENTER
        c.border = _THIN

    # Conditional formatting on LOS column (L5:L12)
    from openpyxl.formatting.rule import CellIsRule
    for los, (bg, fg) in LOS_STYLE.items():
        ws.conditional_formatting.add(
            "L5:L12",
            CellIsRule(
                operator="equal",
                formula=[f'"{los}"'],
                fill=_fill(bg),
                font=_font(bold=True, color=fg),
            )
        )

    # LOS legend below table
    ws.row_dimensions[14].height = 8
    ws.row_dimensions[15].height = 16
    ws.cell(15, 1, "LOS:").font = _font(bold=True)
    ws.cell(15, 1).alignment = _CENTER
    for i, (los, (bg, fg)) in enumerate(LOS_STYLE.items()):
        col = 2 + i
        labels = {"A": "A  ≤10 s", "B": "B  10–20 s", "C": "C  20–35 s",
                  "D": "D  35–55 s", "E": "E  55–80 s", "F": "F  >80 s"}
        c = ws.cell(15, col, labels[los])
        c.font = _font(size=9, bold=True, color=fg)
        c.fill = _fill(bg)
        c.alignment = _CENTER
        c.border = _THIN

    ws.freeze_panes = "C5"


def _build_queue_sheet(wb):
    ws = wb.create_sheet("95th Queue")
    NCOLS = 13

    col_widths = [11, 7, 10, 14, 9, 9, 13, 9, 11, 11, 11, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _sheet_banner(ws,
        "HCM 7th Edition — 95th Percentile Back-of-Queue",
        "Q95 = Nq + 1.65·√Nq     Nq = Nq₁ (uniform) + Nq₂ (overflow)"
        "     Edit parameters in the Input sheet",
        NCOLS)

    hdrs = [
        "Direction", "Period", "Green g\n(s)", "Volume v\n(veh/h)",
        "Lanes n", "u = g/C", "Capacity c\n(veh/h)", "X = v/c",
        "Nq₁\n(veh)", "Nq₂\n(veh)", "Nq avg\n(veh)", "Q95\n(veh)", "Q95\n(m)",
    ]
    ws.row_dimensions[4].height = 32
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 4, col, h, wrap=True)

    for i, (d, period, dcol, green_row, vol_row) in enumerate(DATA_ROWS):
        r = 5 + i
        ws.row_dimensions[r].height = 18
        _common_cols(ws, r, d, period, dcol, green_row, vol_row)

        # I: Nq1 = c*C/3600*(1-u)^2 / (1-min(1,X)*u)
        _calc_cell(ws, r, 9,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"G{r}*Input!$B${R_C}/3600"
            f"*(1-F{r})^2"
            f"/MAX(0.001,1-MIN(1,H{r})*F{r}))",
            fmt="0.0")

        # J: Nq2 = 900*T*((X-1)+sqrt(...))*c/3600
        _calc_cell(ws, r, 10,
            f"=IF(OR(G{r}=0,D{r}=0),0,"
            f"900*Input!$B${R_T}*((H{r}-1)"
            f"+SQRT(MAX(0,(H{r}-1)^2"
            f"+8*Input!$B${R_K}*Input!$B${R_I}*H{r}"
            f"/(G{r}*Input!$B${R_T}))))"
            f"*G{r}/3600)",
            fmt="0.0")

        # K: Nq = max(0, Nq1+Nq2)
        _calc_cell(ws, r, 11, f"=MAX(0,I{r}+J{r})", fmt="0.0")

        # L: Q95 veh
        _calc_cell(ws, r, 12, f"=K{r}+1.65*SQRT(MAX(0,K{r}))", fmt="0.0", bold=True)

        # M: Q95 metres
        _calc_cell(ws, r, 13, f"=L{r}*Input!$B${R_L}", fmt="0", bold=True)

    ws.freeze_panes = "C5"


# ── Build & save ──────────────────────────────────────────────────────────────
def build():
    wb = xl.Workbook()
    _build_input_sheet(wb)
    _build_delay_sheet(wb)
    _build_queue_sheet(wb)

    out_path = "HCM7_Analysis_Tool.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    build()
