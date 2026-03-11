"""
ui_excel.py — Excel ↔ session-state helpers for the JUNC Streamlit UI.

Cell layout (1-indexed, matching Phaser.py read_from_excel):
  Volumes  : morning=row 4, evening=row 5; col = 4 + 4*dir + mov
  Lanes    : row 8;  col = 3 + 8*dir + lane_type_idx
  Nataz    : row 9;  same cols as lanes
  Instructions : col 22 (V), rows 36-46  (11 values)
  Rakal        : col 26 (Z), rows 36-41  (6 values)
  Streets      : row 4,  cols 22-25 (V-Y), one per direction
  Junc info    : col 19 (S), rows 36-40  (5 values, row 38 unused)
"""

from io import BytesIO
import openpyxl as xl

DIRECTIONS = ["N", "S", "E", "W"]
MOVEMENTS  = ["R", "T", "L"]
LANE_TYPES = ["R", "RT", "T", "TL", "L", "RTL", "RL"]

INSTR_KEYS = [
    "capacity", "nlsl", "elwl", "img5", "img6",
    "geo_ns", "geo_ew", "optimize",
    "lrt_orig_ns", "lrt_orig_ew", "inflation",
]
RAKAL_KEYS = ["lrt_enabled", "cycle_time", "lost_time", "headway", "mcu", "gen_lost_time"]
JUNC_KEYS  = ["lrt_line_name", "lrt_ref", None, "proj_counter", "more_info"]


# ---------------------------------------------------------------------------
# Cell coordinate helpers
# ---------------------------------------------------------------------------

def _vol_cell(run: int, dir_idx: int, mov_idx: int) -> tuple[int, int]:
    return (4 + run, 4 + 4 * dir_idx + mov_idx)

def _lane_cell(dir_idx: int, lane_idx: int) -> tuple[int, int]:
    return (8, 3 + 8 * dir_idx + lane_idx)

def _nataz_cell(dir_idx: int, lane_idx: int) -> tuple[int, int]:
    return (9, 3 + 8 * dir_idx + lane_idx)

def _instr_cell(i: int) -> tuple[int, int]:
    return (36 + i, 22)

def _rakal_cell(i: int) -> tuple[int, int]:
    return (36 + i, 26)

def _street_cell(s: int) -> tuple[int, int]:
    return (4, 22 + s)

def _junc_cell(i: int) -> tuple[int, int]:
    return (36 + i, 19)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def default_state() -> dict:
    """Return a blank/default JUNC state dictionary."""
    volumes: dict = {}
    for period in ("morning", "evening"):
        volumes[period] = {d: {m: 0 for m in MOVEMENTS} for d in DIRECTIONS}

    lanes = {d: {lt: 0 for lt in LANE_TYPES} for d in DIRECTIONS}
    for d in DIRECTIONS:
        lanes[d]["T"] = 1        # default: one through-lane per arm

    nataz = {d: {lt: 0 for lt in LANE_TYPES} for d in DIRECTIONS}

    instructions = {
        "capacity":   1800,
        "nlsl":       0,
        "elwl":       0,
        "img5":       0,
        "img6":       0,
        "geo_ns":     3,
        "geo_ew":     3,
        "optimize":   0,
        "lrt_orig_ns": 0,
        "lrt_orig_ew": 0,
        "inflation":  1.0,
    }

    rakal = {
        "lrt_enabled":  0,
        "cycle_time":   120,
        "lost_time":    3,
        "headway":      3,
        "mcu":          1,
        "gen_lost_time": 6,
    }

    streets = {d: "" for d in DIRECTIONS}

    junc = {
        "lrt_line_name": "",
        "lrt_ref":       0,
        "proj_counter":  0,
        "more_info":     "",
    }

    return {
        "volumes":      volumes,
        "lanes":        lanes,
        "nataz":        nataz,
        "instructions": instructions,
        "rakal":        rakal,
        "streets":      streets,
        "junc":         junc,
    }


def _sn(val):
    """Suppress null: return 0 if val is None/falsy-but-not-zero."""
    return val if val is not None else 0


def read_excel_to_state(xlsx_bytes: bytes) -> dict:
    """Read a volume_calculator.xlsx byte-string and return a state dict."""
    wb = xl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    state = default_state()

    # Volumes (morning = run 0, evening = run 1)
    for run_idx, period in enumerate(("morning", "evening")):
        for dir_idx, d in enumerate(DIRECTIONS):
            for mov_idx, m in enumerate(MOVEMENTS):
                r, c = _vol_cell(run_idx, dir_idx, mov_idx)
                val = ws.cell(r, c).value
                state["volumes"][period][d][m] = int(_sn(val) or 0)

    # Lanes & Nataz
    for dir_idx, d in enumerate(DIRECTIONS):
        for lane_idx, lt in enumerate(LANE_TYPES):
            r, c = _lane_cell(dir_idx, lane_idx)
            state["lanes"][d][lt] = int(_sn(ws.cell(r, c).value) or 0)
            r2, c2 = _nataz_cell(dir_idx, lane_idx)
            state["nataz"][d][lt] = int(_sn(ws.cell(r2, c2).value) or 0)

    # Instructions
    for i, key in enumerate(INSTR_KEYS):
        r, c = _instr_cell(i)
        state["instructions"][key] = _sn(ws.cell(r, c).value)
    if not state["instructions"]["inflation"]:
        state["instructions"]["inflation"] = 1.0

    # Rakal
    for i, key in enumerate(RAKAL_KEYS):
        r, c = _rakal_cell(i)
        state["rakal"][key] = _sn(ws.cell(r, c).value)

    # Streets (order: N, S, E, W → cols 22, 23, 24, 25)
    for s, d in enumerate(DIRECTIONS):
        r, c = _street_cell(s)
        val = ws.cell(r, c).value
        state["streets"][d] = str(val) if val else ""

    # Junc instructions
    for i, key in enumerate(JUNC_KEYS):
        if key is None:
            continue
        r, c = _junc_cell(i)
        val = ws.cell(r, c).value
        state["junc"][key] = val if val is not None else (0 if key in ("lrt_ref", "proj_counter") else "")

    wb.close()
    return state


def create_excel_from_state(state: dict, template_bytes: bytes) -> bytes:
    """
    Write *state* into a copy of *template_bytes* (a volume_calculator.xlsx)
    and return the modified workbook as bytes.

    We load with data_only=False so formula cells in the template are preserved;
    only the specific input cells (which are plain-value cells) are overwritten.
    """
    wb = xl.load_workbook(BytesIO(template_bytes))
    ws = wb.active

    # Volumes
    for run_idx, period in enumerate(("morning", "evening")):
        for dir_idx, d in enumerate(DIRECTIONS):
            for mov_idx, m in enumerate(MOVEMENTS):
                r, c = _vol_cell(run_idx, dir_idx, mov_idx)
                ws.cell(r, c).value = int(state["volumes"][period][d][m])

    # Lanes & Nataz
    for dir_idx, d in enumerate(DIRECTIONS):
        for lane_idx, lt in enumerate(LANE_TYPES):
            r, c = _lane_cell(dir_idx, lane_idx)
            ws.cell(r, c).value = int(state["lanes"][d][lt])
            r2, c2 = _nataz_cell(dir_idx, lane_idx)
            ws.cell(r2, c2).value = int(state["nataz"][d][lt])

    # Instructions
    for i, key in enumerate(INSTR_KEYS):
        r, c = _instr_cell(i)
        ws.cell(r, c).value = state["instructions"][key]

    # Rakal
    for i, key in enumerate(RAKAL_KEYS):
        r, c = _rakal_cell(i)
        ws.cell(r, c).value = state["rakal"][key]

    # Streets
    for s, d in enumerate(DIRECTIONS):
        r, c = _street_cell(s)
        ws.cell(r, c).value = state["streets"][d] or None

    # Junc instructions
    for i, key in enumerate(JUNC_KEYS):
        if key is None:
            continue
        r, c = _junc_cell(i)
        ws.cell(r, c).value = state["junc"][key] or None

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
